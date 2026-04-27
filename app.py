import streamlit as st
import json
import re
import os
import io
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv
import google.generativeai as genai

# ── Load API key ──────────────────────────────────────────────────────────────
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY") or st.secrets.get("GEMINI_API_KEY", "")

FIELDS = [
    "transaction_number",
    "transaction_date",
    "entity_name",
    "reason_of_transaction",
    "amount",
    "cgst",
    "sgst",
]

FIELD_LABELS = {
    "transaction_number":    "Transaction No.",
    "transaction_date":      "Transaction Date",
    "entity_name":           "Entity Name",
    "reason_of_transaction": "Reason of Transaction",
    "amount":                "Amount",
    "cgst":                  "CGST",
    "sgst":                  "SGST",
}

PROMPT = """You are an invoice data extraction assistant.
Extract the following fields from this invoice and return ONLY a valid JSON object — no markdown, no explanation.

Fields:
- transaction_number: invoice or bill number
- transaction_date: date on the invoice (YYYY-MM-DD format if possible)
- entity_name: vendor / supplier / company name
- reason_of_transaction: description of goods or services
- amount: total amount as a number only, no currency symbol
- cgst: CGST tax amount as a number only (use 0 if not present)
- sgst: SGST tax amount as a number only (use 0 if not present)

Use empty string "" for any field not found.
Return ONLY the JSON object, nothing else."""

# ── Gemini extraction ─────────────────────────────────────────────────────────
def extract_from_pdf(pdf_bytes: bytes) -> dict:
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-3.1-flash-lite-preview")

    response = model.generate_content([
        PROMPT,
        {"mime_type": "application/pdf", "data": pdf_bytes},
    ])

    content = response.text.strip()
    content = re.sub(r"```(?:json)?|```", "", content).strip()
    match = re.search(r"\{.*\}", content, re.DOTALL)
    if match:
        content = match.group()
    return json.loads(content)

# ── Excel helpers ─────────────────────────────────────────────────────────────
def init_workbook(wb):
    ws = wb.active
    ws.title = "Invoices"
    headers = ["S.No.", "Uploaded At"] + [FIELD_LABELS[f] for f in FIELDS]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    return wb

def create_new_workbook():
    return init_workbook(Workbook())

def load_excel_from_upload(uploaded_file):
    return load_workbook(uploaded_file)

def append_row_to_wb(wb, data):
    ws = wb.active
    row_num = ws.max_row
    row = [row_num, datetime.now().strftime("%Y-%m-%d %H:%M")] + [data.get(f, "") for f in FIELDS]
    ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    return wb

def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def workbook_to_dataframe(wb):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=rows[0])

# ── UI ────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Invoice Extractor", page_icon="🧾", layout="centered")

if not GEMINI_API_KEY:
    st.error("⚠️ Gemini API key not found. Add `GEMINI_API_KEY=...` to your `.env` file (local) or Streamlit secrets (cloud).", icon="🔑")
    st.stop()

# Session state
for key, default in [("workbook", None), ("wb_name", None), ("extracted", None), ("saved", False)]:
    if key not in st.session_state:
        st.session_state[key] = default

st.title("🧾 Invoice Data Extractor")
st.caption("Upload an invoice PDF → extract fields → save to your Excel file.")

# ── STEP 1: Choose Excel file ─────────────────────────────────────────────────
st.subheader("Step 1 · Choose your Excel file")
tab_existing, tab_new = st.tabs(["📂 Open existing file", "✨ Create new file"])

with tab_existing:
    st.write("Upload the Excel file you want to add invoices to.")
    existing_file = st.file_uploader("Select your Excel file (.xlsx)", type=["xlsx"], key="excel_upload", label_visibility="collapsed")
    if existing_file:
        try:
            wb = load_excel_from_upload(existing_file)
            st.session_state.workbook = wb
            st.session_state.wb_name = existing_file.name
            st.session_state.extracted = None
            st.session_state.saved = False
            st.success(f"✅ Loaded **{existing_file.name}** — {wb.active.max_row - 1} existing row(s).")
        except Exception as e:
            st.error(f"Could not open file: {e}")

with tab_new:
    st.write("No existing file? Create a fresh one.")
    new_name = st.text_input("File name", value="invoices.xlsx", placeholder="e.g. Q1-invoices.xlsx")
    if not new_name.endswith(".xlsx"):
        new_name += ".xlsx"
    if st.button("✨ Create new Excel file", use_container_width=True):
        st.session_state.workbook = create_new_workbook()
        st.session_state.wb_name = new_name
        st.session_state.extracted = None
        st.session_state.saved = False
        st.success(f"✅ New file **{new_name}** is ready.")

if st.session_state.workbook:
    st.info(f"📄 Active file: **{st.session_state.wb_name}**")

st.divider()

# ── STEP 2: Upload Invoice PDF ────────────────────────────────────────────────
st.subheader("Step 2 · Upload an invoice PDF")

if not st.session_state.workbook:
    st.warning("Complete Step 1 first — choose or create an Excel file above.")
else:
    uploaded_pdf = st.file_uploader("Select invoice PDF", type=["pdf"], label_visibility="collapsed")

    if uploaded_pdf:
        st.write(f"**{uploaded_pdf.name}** — {uploaded_pdf.size / 1024:.1f} KB")

        if st.button("🔍 Extract Data", type="primary", use_container_width=True):
            st.session_state.extracted = None
            st.session_state.saved = False

            with st.spinner("Parsing PDF and extracting data..."):
                try:
                    result = extract_from_pdf(uploaded_pdf.read())
                    st.session_state.extracted = result
                except json.JSONDecodeError as e:
                    st.error(f"Could not parse response as JSON: {e}")
                except Exception as e:
                    st.error(f"Error: {e}")

    # ── STEP 3: Review & save ─────────────────────────────────────────────────
    if st.session_state.extracted:
        st.divider()
        st.subheader("Step 3 · Review & save")
        st.caption("Check and correct any field before saving.")

        data = st.session_state.extracted
        edited = {}
        col_a, col_b = st.columns(2)

        for i, field in enumerate(FIELDS):
            with (col_a if i % 2 == 0 else col_b):
                edited[field] = st.text_input(
                    FIELD_LABELS[field],
                    value=str(data.get(field, "")),
                    key=f"field_{field}",
                )

        st.divider()
        col_save, col_clear = st.columns([3, 1])
        with col_save:
            if st.button("💾 Save to Excel", type="primary", use_container_width=True):
                wb = append_row_to_wb(st.session_state.workbook, edited)
                st.session_state.workbook = wb
                st.session_state.saved = True
                st.success("✅ Invoice saved!")
        with col_clear:
            if st.button("🗑️ Clear", use_container_width=True):
                st.session_state.extracted = None
                st.session_state.saved = False
                st.rerun()

        if st.session_state.saved:
            st.download_button(
                f"📥 Download {st.session_state.wb_name}",
                data=workbook_to_bytes(st.session_state.workbook),
                file_name=st.session_state.wb_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ── Data preview ──────────────────────────────────────────────────────────────
if st.session_state.workbook:
    df = workbook_to_dataframe(st.session_state.workbook)
    if not df.empty:
        st.divider()
        st.subheader("📊 Invoices in this file")
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.download_button(
            f"📥 Download {st.session_state.wb_name}",
            data=workbook_to_bytes(st.session_state.workbook),
            file_name=st.session_state.wb_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="bottom_download",
        )